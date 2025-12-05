#!/usr/bin/env python3
"""
BPR Unified Runner — Streamlit-ready pipeline

What this version does:
- Exposes a single run_pipeline(...) function you can call from Streamlit.
- Chains your three scripts:
    • BPRproofing.main()  (runs on the PDF + expected order CSV/Excel)
    • proofing.process_pdf()  (Profiles + optional TOC Review)
    • newvalidate.run_checks() (Profiles vs BBB; returns validated Profiles + errors)
- Consolidates results into a single Excel workbook with multiple tabs:
    • Profiles (validated)
    • (optional) Profiles_Raw
    • TOC Review (if proofing returns it)
    • Selected sheets copied from BPRproofing output
    • Errors (combined from BPRproofing, newvalidate, and Profiles sheet synthesis)

This version:
- DOES NOT use Tkinter or any GUI dialogs.
- DOES NOT prompt for input paths or save paths.
- DOES NOT have a __main__ block.
- Instead, you must pass explicit file paths to run_pipeline(), which makes it easy
  to call from a Streamlit app that handles uploads and downloads.
"""

from __future__ import annotations

import os
import re
import traceback
from typing import Optional, List, Tuple, Dict, Any

import pandas as pd

# --- Local modules (your scripts) ---
import sl_bprproofing as bpr
import sl_proofing as prof
import sl_newvalidate as nv

# -----------------------------
# Configuration
# -----------------------------
USE_ONE_REF_FILE = True  # kept for compatibility; you may ignore in Streamlit

# Keep only one Profiles tab, named exactly like the separate runs
PROFILES_VALIDATED_NAME = "Profiles"   # write checked_df here

# Optionally also keep the raw Profiles from proofing
KEEP_RAW_PROFILES = False
RAW_PROFILES_NAME = "Profiles_Raw"

# -----------------------------
# Step 1 — Run BPRproofing *in-process* with monkeypatched pickers
# -----------------------------
def run_bprproofing_inprocess(pdf_path: str, bpr_csv_path: str) -> Optional[str]:
    """
    Call BPRproofing.main() while forcing its dialogs to return our chosen paths.
    Returns the final workbook path BPRproofing produced (best-effort).

    In Streamlit, you should:
    - Save the uploaded PDF to pdf_path.
    - Save the uploaded expected-order Excel/CSV to bpr_csv_path.
    - Then call this function.
    """
    print("\n— Step 1/3: Running BPRproofing with provided paths —")

    # Save originals to restore later
    orig_choose_file = getattr(bpr, "choose_file_dialog", None)
    orig_choose_csv  = getattr(bpr, "choose_csv_dialog", None)

    try:
        # Monkeypatch the dialog helpers to avoid GUI in Streamlit
        def _return_pdf():
            print(f"[BPR] Using provided PDF: {pdf_path}")
            return pdf_path

        def _return_csv():
            print(f"[BPR] Using provided CSV/Excel: {bpr_csv_path}")
            return bpr_csv_path

        if orig_choose_file is None or orig_choose_csv is None:
            print("[WARN] BPRproofing does not expose choose_file_dialog / choose_csv_dialog; running main() anyway.")
        else:
            bpr.choose_file_dialog = _return_pdf  # type: ignore[attr-defined]
            bpr.choose_csv_dialog  = _return_csv  # type: ignore[attr-defined]

        # Run main(); it writes the workbook near the PDF and returns None.
        bpr.main()

        # Heuristic: BPRproofing writes to <pdf_stem>_tocsplit.xlsx at the end.
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        candidate = os.path.join(os.path.dirname(pdf_path), f"{base}_tocsplit.xlsx")
        if os.path.isfile(candidate):
            print(f"[BPR] Detected BPRproofing output: {candidate}")
            return candidate
        # fallback to <base>.xlsx
        alt = os.path.join(os.path.dirname(pdf_path), f"{base}.xlsx")
        return alt if os.path.isfile(alt) else None

    finally:
        # Restore originals
        if orig_choose_file is not None:
            bpr.choose_file_dialog = orig_choose_file  # type: ignore[attr-defined]
        if orig_choose_csv is not None:
            bpr.choose_csv_dialog = orig_choose_csv    # type: ignore[attr-defined]


# -----------------------------
# Step 2 — Proofing + NewValidate (programmatic)
# -----------------------------
class PathLike(str):
    @property
    def stem(self):
        return os.path.splitext(os.path.basename(self))[0]

def run_proofing_and_validate(
    pdf_path: str,
    bbb_path: str
) -> tuple[pd.DataFrame, pd.DataFrame, Optional[pd.DataFrame], Optional[pd.DataFrame]]:
    """
    Returns: (profiles_df_raw, profiles_df_validated, toc_review_df or None, nv_errors_df or None)
    - profiles_df_raw: from proofing.process_pdf
    - profiles_df_validated: from newvalidate.run_checks
    - toc_review_df: if proofing provides it (tuple/dict return), else None
    - nv_errors_df: errors DataFrame returned/exposed by newvalidate if available

    Streamlit will pass:
    - pdf_path: path to a temp file containing the uploaded PDF.
    - bbb_path: path to a temp file containing the uploaded BBB reference Excel/CSV.
    """
    print("\n— Step 2/3: Extracting Profiles and Running Validation —")

    # ---- Profiles & TOC Review (proofing) ----
    print("Extracting profiles from PDF… (proofing.process_pdf)")
    toc_review_df: Optional[pd.DataFrame] = None
    profiles_out = prof.process_pdf(PathLike(pdf_path))

    # Flexibly accept multiple shapes from proofing.process_pdf
    if isinstance(profiles_out, tuple):
        profiles_df = profiles_out[0]
        if len(profiles_out) > 1 and isinstance(profiles_out[1], pd.DataFrame):
            toc_review_df = profiles_out[1]
    elif isinstance(profiles_out, dict):
        profiles_df = profiles_out.get("Profiles") or next(
            (v for k, v in profiles_out.items() if isinstance(v, pd.DataFrame)),
            None
        )
        if "TOC Review" in profiles_out and isinstance(profiles_out["TOC Review"], pd.DataFrame):
            toc_review_df = profiles_out["TOC Review"]
        if profiles_df is None:
            raise RuntimeError("proofing.process_pdf did not return a Profiles DataFrame.")
    else:
        profiles_df = profiles_out

    if not isinstance(profiles_df, pd.DataFrame):
        raise RuntimeError("proofing.process_pdf did not yield a DataFrame for Profiles.")

    # ---- BBB table ----
    print(f"Loading BBB from: {bbb_path}")
    bbb_df = nv.load_table(bbb_path)

    # ---- Validation (newvalidate) ----
    print("Running newvalidate.run_checks (Profiles vs BBB)…")
    checked_df: pd.DataFrame
    nv_errors_df: Optional[pd.DataFrame] = None

    try:
        result = nv.run_checks(profiles_df.copy(), bbb_df)
        if isinstance(result, tuple):
            # common: (checked_df, errors_df)
            checked_df = result[0]
            if len(result) >= 2 and isinstance(result[1], pd.DataFrame):
                nv_errors_df = result[1]
        elif isinstance(result, dict):
            checked_df = (
                result.get("profiles")
                or result.get("validated")
                or result.get("checked")       # type: ignore[assignment]
            )
            nv_errors_df = result.get("errors") if isinstance(result.get("errors"), pd.DataFrame) else None
            if not isinstance(checked_df, pd.DataFrame):
                raise RuntimeError("newvalidate.run_checks did not return a DataFrame for Profiles.")
        else:
            checked_df = result  # assume plain DataFrame
            # try to locate a module-level errors df if newvalidate exposes one
            for attr in ("errors_df", "ERRORS_DF", "last_errors_df", "errors"):
                if hasattr(nv, attr) and isinstance(getattr(nv, attr), pd.DataFrame):
                    nv_errors_df = getattr(nv, attr)
                    break
    except Exception as e:
        print(f"[WARN] newvalidate.run_checks returned unexpected shape: {e}")
        checked_df = nv.run_checks(profiles_df.copy(), bbb_df)  # try again, assume DF
        if not isinstance(checked_df, pd.DataFrame):
            raise

    return profiles_df, checked_df, toc_review_df, nv_errors_df


# -----------------------------
# Utilities for Excel writing (no Tables)
# -----------------------------
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def _excel_autowidth(ws, max_width=60):
    """Autosize columns based on first line of content; cap width."""
    for col in ws.columns:
        first_cell = next((c for c in col if c is not None), None)
        if first_cell is None:
            continue
        try:
            letter = get_column_letter(
                getattr(first_cell, "column", getattr(first_cell, "col_idx", 1))
            )
        except Exception:
            continue
        maxlen = 0
        for cell in col:
            if cell is None:
                continue
            val = "" if cell.value is None else str(cell.value)
            first_line = val.split("\n", 1)[0]
            maxlen = max(maxlen, len(first_line) + (3 if "\n" in val else 0))
        ws.column_dimensions[letter].width = min(maxlen + 2, max_width)

def _write_df_sheet(wb, sheet_name, df, *, add_header_filter: bool = False):
    """Create/replace sheet, write DataFrame, style header, freeze top row,
       and optionally add a simple header AutoFilter (NO Excel Table)."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Header
    hdr_font = Font(bold=True)
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = hdr_font
        c.alignment = Alignment(vertical="top", wrap_text=True)

    # Data
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    # Add filter only if requested and there is at least one data row and col
    if add_header_filter and df.shape[0] >= 1 and df.shape[1] >= 1:
        last_col = get_column_letter(df.shape[1])
        last_row = df.shape[0] + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    _excel_autowidth(ws)
    return ws

def copy_sheet_if_exists(src_xlsx: str, writer: pd.ExcelWriter, sheet_name: str, out_tab: Optional[str] = None):
    try:
        xl = pd.ExcelFile(src_xlsx, engine="openpyxl")
        if sheet_name in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet_name)
            df.to_excel(writer, index=False, sheet_name=out_tab or sheet_name)
            print(f"  ✓ Copied sheet: {sheet_name}")
        else:
            print(f"  • Sheet not found (skipped): {sheet_name}")
    except Exception as e:
        print(f"  ! Could not copy '{sheet_name}': {e}")

def _gather_errors_from_workbook(xlsx_path: str) -> List[pd.DataFrame]:
    """Return list of DataFrames for any sheet with 'error' in its name."""
    outs: List[pd.DataFrame] = []
    try:
        xl = pd.ExcelFile(xlsx_path, engine="openpyxl")
        for name in xl.sheet_names:
            if "error" in name.lower():
                try:
                    df = pd.read_excel(xl, sheet_name=name)
                    outs.append(df)
                except Exception:
                    pass
    except Exception:
        pass
    return outs

def _dedupe_errors(df: pd.DataFrame) -> pd.DataFrame:
    """Light dedupe by common subset."""
    for col in ["Sheet","Key","Issue","Expected","Found","Page"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    subset = [c for c in ["Sheet","Row","Key","Issue","Page"] if c in df.columns]
    if subset:
        return df.drop_duplicates(subset=subset, keep="first").reset_index(drop=True)
    return df


# -----------------------------
# Text normalization & cleaning
# -----------------------------
_EMPTY_TOKENS = {"", "nan", "none", "null", "n/a", "na", "-"}

def _norm_text(v: object) -> str:
    s = ("" if v is None else str(v)).strip()
    return "" if s.lower() in _EMPTY_TOKENS else s

def _clean_issue_text(issue: object) -> str:
    """Remove 'nan'-style placeholders and empty labeled lines."""
    s = ("" if issue is None else str(issue))
    # Collapse whitespace/newlines
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n+", "\n", s).strip()

    # Drop lines that are just placeholders or label+placeholder
    lines = []
    for line in s.split("\n"):
        line = line.strip()
        if line.lower() in _EMPTY_TOKENS:
            continue
        if re.fullmatch(r"\[(Internal|Compare)\]\s*(nan|none|null|n/?a|-)?\s*", line, flags=re.I):
            continue
        if re.fullmatch(r"\[(Internal|Compare)\]\s*", line, flags=re.I):
            continue
        lines.append(line)

    s = "\n".join(lines).strip()
    return s


# -----------------------------
# Build Errors from Profiles
# -----------------------------
def _errors_from_profiles_sheet(
    xlsx_path: str,
    sheet_name: str = "Profiles"
) -> Optional[pd.DataFrame]:
    """
    Read 'Profiles', synthesize Errors rows to:
    Sheet, Row, Key, Issue, Expected, Found, Page
    Prefers prebuilt 'ERRORS' column if present; else combines Notes_Internal + Notes_Compare.
    """
    try:
        xl = pd.ExcelFile(xlsx_path, engine="openpyxl")
        if sheet_name not in xl.sheet_names:
            return None
        df = pd.read_excel(xl, sheet_name=sheet_name)

        # Normalize headers
        df.columns = [str(c).strip() for c in df.columns]

        def find_col(cands):
            lcands = [x.lower().strip() for x in cands]
            for c in df.columns:
                if c.lower().strip() in lcands:
                    return c
            return None

        # Prefer a ready-made ERRORS column if present
        col_errors = find_col(["ERRORS","Errors","Error","Issue","Issues"])

        col_notes_i = find_col(["Notes_Internal","Notes Internal","Internal Notes","Notes (Internal)","Internal"])
        col_notes_c = find_col(["Notes_Compare","Notes Compare","Compare Notes","Notes (Compare)","Compare"])
        col_category = find_col(["Category","Cat","Category Name"])
        col_pubnum  = find_col([
            "Published Name + Number","Published Name+Number","Published Name Number",
            "Published Name/Number","Published Name & Number","Published Name and Number",
            "Published Name","Name + Number","Name+Number"
        ])
        col_page    = find_col(["Page","Pg","Page #","Page Number","PageNo"])

        # Build Issue text
        if col_errors:
            issues = df[col_errors].map(_clean_issue_text)
        else:
            def build_issue(row):
                parts = []
                if col_notes_i:
                    v = _norm_text(row.get(col_notes_i, ""))
                    if v:
                        parts.append(f"[Internal] {v}")
                if col_notes_c:
                    v = _norm_text(row.get(col_notes_c, ""))
                    if v:
                        parts.append(f"[Compare] {v}")
                return _clean_issue_text("\n".join(parts))
            issues = df.apply(build_issue, axis=1)

        # Keep only non-empty issues
        mask = issues.astype(str).str.strip().astype(bool)
        if not mask.any():
            return None

        excel_rows = (df.index + 2).astype(int)

        def build_key(row):
            left  = _norm_text(row.get(col_category, "")) if col_category else ""
            right = _norm_text(row.get(col_pubnum , "")) if col_pubnum  else ""
            return f"{left} / {right}".strip(" /")

        out = pd.DataFrame({
            "Sheet": "Profiles",
            "Row": excel_rows,
            "Key": df.apply(build_key, axis=1),
            "Issue": issues,
            "Expected": "",
            "Found": "",
            "Page": df[col_page] if col_page in df.columns else "",
        })

        return out[mask].reset_index(drop=True)

    except Exception:
        return None


# -----------------------------
# Step 3 — Consolidate into one Excel
# -----------------------------
def consolidate_to_single_excel(
    save_path: str,
    profiles_df: pd.DataFrame,               # raw profiles (optional)
    checked_df: pd.DataFrame,                # validated → Profiles
    bpr_workbook_path: Optional[str],
    toc_review_df: Optional[pd.DataFrame],   # from proofing if returned
    nv_errors_df: Optional[pd.DataFrame],    # <-- pass-through from run_proofing_and_validate
):
    """
    Writes the combined workbook to save_path.

    Streamlit will typically:
    - Provide save_path as a temporary file.
    - Then read that file back into bytes to send as a download.
    """
    print(f"\n— Step 3/3: Writing combined workbook —\n{save_path}")

    # 1) Write with pandas first
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        if KEEP_RAW_PROFILES:
            profiles_df.to_excel(writer, index=False, sheet_name=RAW_PROFILES_NAME)

        # Validated becomes the single Profiles tab
        checked_df.to_excel(writer, index=False, sheet_name=PROFILES_VALIDATED_NAME)

        # Write TOC Review if we got it from proofing
        if isinstance(toc_review_df, pd.DataFrame) and not toc_review_df.empty:
            toc_review_df.to_excel(writer, index=False, sheet_name="TOC Review")

        # Copy helpful tabs from BPRproofing workbook (except Errors — handled later)
        if bpr_workbook_path and os.path.isfile(bpr_workbook_path):
            print("Copying sheets from BPRproofing workbook (if present):")
            for tab in ["Listings_Split", "TOC Review"]:
                copy_sheet_if_exists(bpr_workbook_path, writer, tab)

    # 2) Build Errors = BPR Errors + newvalidate Errors + Profiles-synth (always include)
    wb = load_workbook(save_path)

    errors_sources: List[pd.DataFrame] = []

    def _attach_source(df: pd.DataFrame, label: str) -> Optional[pd.DataFrame]:
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            return None
        out = df.copy()
        if "Source" not in out.columns:
            out["Source"] = label
        return out

    # a) Errors from BPRproofing workbook
    if bpr_workbook_path and os.path.isfile(bpr_workbook_path):
        for d in _gather_errors_from_workbook(bpr_workbook_path):
            tagged = _attach_source(d, "BPRproofing")
            if tagged is not None:
                errors_sources.append(tagged)

    # b) Errors from newvalidate (explicit pass-through)
    tagged_nv = _attach_source(nv_errors_df, "newvalidate")
    if tagged_nv is not None:
        errors_sources.append(tagged_nv)

    # c) Profiles-synth (ALWAYS include if any issues found)
    synth = _errors_from_profiles_sheet(save_path, sheet_name=PROFILES_VALIDATED_NAME)
    tagged_synth = _attach_source(synth, "Profiles")
    if tagged_synth is not None:
        errors_sources.append(tagged_synth)

    # d) Last-chance: module-level accessors (if your newvalidate exposes globals)
    if tagged_nv is None:
        fallback_nv = _attach_source(_try_get_newvalidate_errors_df(), "newvalidate (module)")
        if fallback_nv is not None:
            errors_sources.append(fallback_nv)

    # Merge, align, dedupe
    if errors_sources:
        # Union of columns
        all_cols: List[str] = []
        for d in errors_sources:
            for c in d.columns:
                if c not in all_cols:
                    all_cols.append(c)

        combined = pd.concat(
            [d.reindex(columns=all_cols) for d in errors_sources],
            ignore_index=True
        )

        # Clean up placeholder issues across all sources, then drop empties
        if "Issue" in combined.columns:
            combined["Issue"] = combined["Issue"].map(_clean_issue_text)
            combined = combined[
                combined["Issue"].astype(str).str.strip().astype(bool)
            ].reset_index(drop=True)

        # Normalize to desired schema ordering if present
        desired = ["Sheet","Row","Key","Issue","Expected","Found","Page","Source"]
        ordered = [c for c in desired if c in combined.columns] + [
            c for c in combined.columns if c not in desired
        ]
        combined = combined.reindex(columns=ordered)
        combined = _dedupe_errors(combined)

        # Debug counts
        try:
            counts = combined["Source"].value_counts(dropna=False).to_dict()
            print(f"  • Errors rows by source: {counts}")
        except Exception:
            pass
    else:
        combined = pd.DataFrame(
            columns=["Sheet","Row","Key","Issue","Expected","Found","Page","Source"]
        )

    # Only the Errors tab gets a simple header filter; no Excel Tables anywhere
    _write_df_sheet(wb, "Errors", combined, add_header_filter=True)
    print("  ✓ Wrote 'Errors' tab (combined).")

    wb.save(save_path)
    wb.close()


# -----------------------------
# Pull errors DF from newvalidate if exposed
# -----------------------------
def _try_get_newvalidate_errors_df() -> Optional[pd.DataFrame]:
    """
    Best-effort: pull an Errors DataFrame produced by newvalidate when run in-process.
    Supports several common patterns without requiring a separate file.
    """
    # Common module-level attributes
    for attr in ("errors_df", "ERRORS_DF", "last_errors_df", "errors"):
        if hasattr(nv, attr):
            val = getattr(nv, attr)
            if isinstance(val, pd.DataFrame):
                return val

    # Common accessor functions
    for fn_name in ("get_errors_df", "errors_dataframe", "build_errors_df"):
        if hasattr(nv, fn_name):
            try:
                val = getattr(nv, fn_name)()
                if isinstance(val, pd.DataFrame):
                    return val
            except Exception:
                pass

    # Nothing found
    return None


# -----------------------------
# Final tidy (keep Profiles and Errors)
# -----------------------------
def tidy_up_final_workbook(save_path: str):
    """
    Keep Profiles; optionally remove Profiles_Raw; rename other tabs if desired.
    """
    wb = load_workbook(save_path)

    # Remove raw if not keeping it
    if not KEEP_RAW_PROFILES and RAW_PROFILES_NAME in wb.sheetnames:
        del wb[RAW_PROFILES_NAME]
        print(f"🗑️  Deleted '{RAW_PROFILES_NAME}' (kept '{PROFILES_VALIDATED_NAME}').")

    # Optional renames
    rename_map = {
        "Listings_Split": "Cat Listings",
        "TOC Review": "TOC",
    }
    for old_name, new_name in rename_map.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name
            print(f"✏️  Renamed '{old_name}' → '{new_name}'")

    wb.save(save_path)
    wb.close()
    print("🧹 Final workbook tidy-up complete.")


# -----------------------------
# Public API for Streamlit
# -----------------------------
def run_pipeline(
    pdf_path: str,
    bpr_csv_path: str,
    bbb_path: str,
    save_path: str,
) -> str:
    """
    Streamlit-friendly orchestration function.

    Parameters
    ----------
    pdf_path : str
        Path to the uploaded Best Pick PDF (saved to disk by Streamlit).
    bpr_csv_path : str
        Path to the expected-order Excel/CSV used by BPRproofing.
    bbb_path : str
        Path to the BBB reference Excel/CSV used by newvalidate.
    save_path : str
        Path where the combined Excel workbook should be written.

    Returns
    -------
    str
        The same save_path, for convenience.

    Typical Streamlit usage:
    - Save uploads to temp files (one PDF, one BPR ref, one BBB ref).
    - Call run_pipeline(pdf_path, bpr_csv_path, bbb_path, save_path).
    - Read save_path back into bytes and offer as a download.
    """
    # 1) Run BPRproofing with the same selections
    bpr_out_xlsx = run_bprproofing_inprocess(pdf_path, bpr_csv_path)

    # 2) Profiles + Validation using the same PDF and BBB Excel/CSV
    try:
        profiles_df, checked_df, toc_review_df, nv_errors_df = run_proofing_and_validate(
            pdf_path,
            bbb_path,
        )
    except Exception as e:
        print(f"[ERROR] Failed during Profiles/Validate stage: {e}")
        traceback.print_exc()
        raise

    # 3) Save combined workbook
    try:
        consolidate_to_single_excel(
            save_path,
            profiles_df,
            checked_df,
            bpr_out_xlsx,
            toc_review_df,
            nv_errors_df,
        )
        tidy_up_final_workbook(save_path)
    except Exception as e:
        print(f"[ERROR] Failed to write combined workbook: {e}")
        traceback.print_exc()
        raise

    return save_path
