import os
import tempfile
import pandas as pd
from PyPDF2 import PdfReader
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from typing import List, Optional, Tuple, Dict, Any
import json

PHRASE = "Best Pick Reports recommends:"         # for Listings (case-insensitive)
TEXT_COL_CANDIDATES = ["text", "Text", "TEXT"]   # preferred column(s) to search

def extract_pdf_text(pdf_path: str):
    """Extract text from all pages of the PDF."""
    reader = PdfReader(pdf_path)
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append(text)
        print(f"[INFO] Extracted page {i}")
    return pages

def save_to_excel(pdf_path: str, pages: list) -> str:
    """Save the extracted text to Excel with the same base name as the PDF."""
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    out_excel = os.path.join(os.path.dirname(pdf_path), f"{base}.xlsx")
    combined = "\n\n".join(pages)

    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        pd.DataFrame({"page": range(1, len(pages)+1), "text": pages}).to_excel(
            writer, index=False, sheet_name="Pages"
        )
        pd.DataFrame([{
            "pdf": pdf_path,
            "pages": len(pages),
            "characters_total": len(combined),
            "method": "PyPDF2 (no OCR)"
        }]).to_excel(writer, index=False, sheet_name="Summary")

    print(f"[OK] Saved Excel: {out_excel}")
    return out_excel

# ----------------- Parse helpers -----------------
def pick_text_series(df: pd.DataFrame) -> Tuple[pd.Series, str]:
    """Pick a column to search Listings phrase WITHOUT modifying df."""
    for col in TEXT_COL_CANDIDATES:
        if col in df.columns:
            return df[col].fillna("").astype(str), col
    obj_cols = [c for c in df.columns if df[c].dtype == "object"]
    if obj_cols:
        return df[obj_cols].fillna("").astype(str).agg(" | ".join, axis=1), " | ".join(obj_cols)
    return df.astype(str).agg(" | ".join, axis=1), "<all columns stringified>"

def build_tabs_keep_rows(df_pages: pd.DataFrame):
    """
    Keep ENTIRE original rows:
      - TOC: rows with page containing "Table of Contents" + last row from final page
      - Listings: rows where chosen text contains PHRASE (case-insensitive)
      - Profiles: all other rows
    """
    if "page" not in df_pages.columns:
        raise ValueError("Expected 'Pages' sheet to have a 'page' column (1-based).")

    df = df_pages.copy()
    df["_row_id__"] = range(len(df))
    page_num = pd.to_numeric(df["page"], errors="coerce")

    # TOC: Find page(s) containing "Table of Contents"
    toc_parts = []
    text_series, _ = pick_text_series(df)
    toc_pattern = re.compile(r'table\s+of\s+contents', re.IGNORECASE)
    toc_mask = text_series.apply(lambda s: bool(toc_pattern.search(s)))
    if toc_mask.any():
        toc_parts.append(df.loc[toc_mask])
        toc_page_nums = page_num.loc[toc_mask].tolist()
        print(f"[INFO] Found 'Table of Contents' on page(s): {toc_page_nums}")
    else:
        print("[WARN] 'Table of Contents' not found in any page; TOC sheet may be empty.")
    
    # Also include last row from final page (back cover TOC)
    if page_num.notna().any():
        last_page = int(page_num.max())
        last_rows = df.loc[page_num == last_page]
        if not last_rows.empty:
            toc_parts.append(last_rows.tail(1))
    df_toc = pd.concat(toc_parts, ignore_index=False) if toc_parts else df.iloc[0:0]
    df_toc = df_toc.copy()

    # Listings
    text_series, used_col = pick_text_series(df)
    pat = re.compile(re.escape(PHRASE), re.IGNORECASE)
    listings_mask = text_series.apply(lambda s: bool(pat.search(s)))
    df_listings = df.loc[listings_mask].copy()

    # Profiles
    exclude_ids = set(df_toc["_row_id__"].tolist()) | set(df_listings["_row_id__"].tolist())
    df_profiles = df.loc[~df["_row_id__"].isin(exclude_ids)].copy()

    for d in (df_toc, df_listings, df_profiles):
        if "_row_id__" in d.columns:
            d.drop(columns=["_row_id__"], inplace=True)

    return df_toc, df_listings, df_profiles, used_col

# Categories to ignore for "No matching category in Listings_Split" checks
IGNORE_TOC_ONLY_CATEGORIES = {
    "air conditioning & heating",
    "impact windows & doors",
    "additional information",
    "faqs",
    "quick reference sheet",
    "homeowner protection tips",
    "seasonal maintenance checklist",
}

def _derive_company_from_ratings_text(s: str) -> str | None:
    """
    Ratings Table looks like:
      "R.S. Andrews Services, Inc. 770-663-5440\n13th year as a Best Pick"
      "Estes Services, Heat, Air, Plumbing & Electrical 770-277-7990\n13th year as a Best Pick"
    Heuristics:
      - Use the first non-empty line.
      - Strip trailing phone number (various formats).
      - Keep commas/dashes that are part of the company name.
    """
    if not s:
        return None
    txt = str(s).strip().replace("\r\n", "\n").replace("\r", "\n")
    first_line = next((ln.strip() for ln in txt.split("\n") if ln.strip()), "")
    if not first_line:
        return None

    first_line = re.sub(r'\s+as\s+a\s+best\s+pick.*$', '', first_line, flags=re.I).strip()

    phone_pat = re.compile(r'\s*(\(?\d{3}\)?[\s\-.]*\d{3}[\s\-.]*\d{4})\s*$', flags=re.I)
    first_line = phone_pat.sub('', first_line).strip()

    first_line = first_line.lstrip("■•·").strip()

    company = re.sub(r'\s+', ' ', first_line).strip()
    return company or None

def write_into_existing_workbook(src_path: str,
                                 df_toc: pd.DataFrame,
                                 df_listings: pd.DataFrame,
                                 df_profiles: pd.DataFrame) -> str:
    """
    Write/replace TOC, Listings, Profiles INTO the same workbook (src_path).
    """
    try:
        with pd.ExcelWriter(src_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_toc.to_excel(writer, index=False, sheet_name="TOC")
            df_listings.to_excel(writer, index=False, sheet_name="Listings")
            df_profiles.to_excel(writer, index=False, sheet_name="Profiles")
    except PermissionError:
        raise PermissionError("Close the workbook in Excel and try again.")

    # Enable wrap_text on cells containing line breaks
    wb = load_workbook(src_path)
    for ws_name in ["TOC", "Listings", "Profiles"]:
        if ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "\n" in cell.value:
                        cell.alignment = Alignment(wrap_text=True)
    wb.save(src_path)
    return src_path

def normalize_text(s: Optional[str]) -> str:
    if not s:
        return ""
    return (s.replace("\xa0", " ")
             .replace("\u2002", " ")
             .replace("\u2003", " ")
             .replace("\u2026", "...")
             .replace("\t", " ")
             .strip())

def strip_before_toc(text: str) -> str:
    """Remove everything before and including 'Table of Contents' (case-insensitive)."""
    if not text:
        return ""
    m = re.search(r'table\s+of\s+contents', text, flags=re.IGNORECASE)
    return text[m.end():].lstrip() if m else text

def parse_pairs_split_on_numbers(block: str) -> List[Tuple[str, int]]:
    """
    Scan the entire block and split whenever a NUMBER ends.
    Only accept numbers that look like TOC page refs:
      - preceded nearby (within ~40 chars) by dot leaders / dashes / or >=2 spaces.
    """
    if not block:
        return []
    txt = normalize_text(block)

    txt = re.sub(r'[•·]', '.', txt)
    txt = re.sub(r'\.{2,}', '  ', txt)   # dot leaders -> double space
    txt = re.sub(r'[–—]+', '-', txt)

    pairs: List[Tuple[str, int]] = []
    last_end = 0

    for m in re.finditer(r'(\d{1,5})', txt):
        num = int(m.group(1))
        window_start = max(0, m.start() - 40)
        before = txt[window_start:m.start()]
        if not re.search(r'(?:\s{2,}|-+)', before):
            continue  # not a TOC-like number

        cat_chunk = txt[last_end:m.start()]
        cat = re.sub(r'[\s\.\-]+$', '', cat_chunk).strip()
        if cat:
            pairs.append((cat, num))
            last_end = m.end()
        else:
            last_end = m.end()
    return pairs

def write_split_sheet(wb, front_pairs: List[Tuple[str,int]], back_pairs: List[Tuple[str,int]]):
    """Create/overwrite TOC Review with columns: Front TOC | Front TOC # | Back TOC | Back TOC #"""
    if "TOC Review" in wb.sheetnames:
        del wb["TOC Review"]
    ws = wb.create_sheet("TOC Review")

    headers = ["Front TOC", "Front TOC #", "Back TOC", "Back TOC #"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    n = max(len(front_pairs), len(back_pairs))
    for i in range(n):
        if i < len(front_pairs):
            ws.cell(row=i+2, column=1, value=front_pairs[i][0])
            ws.cell(row=i+2, column=2, value=front_pairs[i][1])
        if i < len(back_pairs):
            ws.cell(row=i+2, column=3, value=back_pairs[i][0])
            ws.cell(row=i+2, column=4, value=back_pairs[i][1])

    ws.column_dimensions[get_column_letter(1)].width = 40
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 40
    ws.column_dimensions[get_column_letter(4)].width = 12

def norm(s: Optional[str]) -> str:
    if s is None:
        return ""
    return (str(s)
            .replace("\xa0", " ")
            .replace("\u2002", " ")
            .replace("\u2003", " ")
            .replace("\u2026", "...")
            .replace("\t", " "))

def extract_page_from_text(text: str) -> Optional[int]:
    m = re.search(r"^\s*(\d{1,5})\b", text)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            return None
    for line in text.splitlines():
        if line.strip():
            m2 = re.match(r"^\s*(\d{1,5})\b", line)
            if m2:
                try:
                    return int(m2.group(1))
                except ValueError:
                    return None
            break
    return None

def slice_between(text: str, start_pat: str, end_pat: str) -> str:
    if not text:
        return ""
    m = re.search(start_pat, text, re.I | re.S)
    if not m:
        return ""
    start = m.end()
    m2 = re.search(end_pat, text[start:], re.I | re.S)
    if not m2:
        return ""
    return text[start:start + m2.start()].strip()

def slice_after(text: str, start_pat: str) -> str:
    if not text:
        return ""
    m = re.search(start_pat, text, re.I | re.S)
    if not m:
        return ""
    return text[m.end():].strip()

def content_from_tail(tail: str) -> str:
    if not tail:
        return ""
    m = re.search(r"\bCommon\b", tail, re.I)
    return tail[m.start():].strip() if m else ""

def clean_category(s: str) -> str:
    return " ".join(s.split()).strip(": ").strip()

def clean_license_expl(s: str) -> str:
    return s.strip()

def ratings_entries_from_tail(tail: str) -> List[str]:
    """Entries start after '■' and end at 'as a Best Pick' (inclusive)."""
    if not tail:
        return []
    txt = norm(tail)
    pat = re.compile(r"■\s*(.+?as\s+a\s+Best\s+Pick)", re.I | re.S)
    return [m.group(1).strip() for m in pat.finditer(txt)]

def process_listing_row(row: pd.Series) -> List[Dict[str, Any]]:
    """
    From one Listings row (needs 'text'; 'page' optional), return rows duplicated per Ratings entry.
    """
    original: Dict[str, Any] = row.to_dict()
    text = norm(original.get("text", ""))

    page_from_text = extract_page_from_text(text)
    if page_from_text is not None:
        original["page"] = page_from_text

    category_raw = slice_between(
        text, r"Best\s*Pick\s*Reports\s*recommends\s*:\s*", r"Trade\s*License\s*Information"
    )
    category = clean_category(category_raw)

    license_expl = slice_between(
        text, r"Trade\s*License\s*Information", r"Scan\s*for\s*additional\s*educational\s*content"
    )
    license_expl = clean_license_expl(license_expl)

    tail = slice_after(text, r"Scan\s*for\s*additional\s*educational\s*content")

    ratings_entries = ratings_entries_from_tail(tail)
    content_block = content_from_tail(tail)

    out_rows: List[Dict[str, Any]] = []
    if ratings_entries:
        for i, entry in enumerate(ratings_entries, start=1):
            r = dict(original)
            r["Category"] = category
            r["License Explanation"] = license_expl
            r["Ratings Table"] = entry
            r["Company Order"] = i
            r["Content"] = content_block
            out_rows.append(r)
    else:
        r = dict(original)
        r["Category"] = category
        r["License Explanation"] = license_expl
        r["Ratings Table"] = ""
        r["Company Order"] = None
        r["Content"] = content_block
        out_rows.append(r)

    return out_rows

def delete_sheets(src_path: str, sheet_names_to_delete: list[str]) -> None:
    """Delete the given sheet names from an Excel workbook and save."""
    wb = load_workbook(src_path)
    for name in sheet_names_to_delete:
        if name in wb.sheetnames:
            del wb[name]
            print(f"[INFO] Deleted sheet: {name}")
    wb.save(src_path)

def first_empty_col(ws) -> int:
    """Return the first empty (append) column index."""
    return (ws.max_column or 0) + 1

def _format_errors_ws(ws):
    """Freeze header, set column widths, and wrap long text on the Errors sheet."""
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = "A2"

    width_map = {
        1: 14,  # Sheet
        2: 8,   # Row
        3: 60,  # Key
        4: 50,  # Issue
        5: 14,  # Expected
        6: 14,  # Found
        7: 10,  # Page
    }
    for idx, w in width_map.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    wrap_cols = {3, 4, 5, 6}
    for r in range(1, ws.max_row + 1):
        for c in wrap_cols:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

def append_error(errors: list, sheet: str, row: int, key: str, issue: str,
                 expected: str | int | None = None, found: str | int | None = None,
                 page: Optional[int] = None):
    errors.append({
        "Sheet": sheet,
        "Row": row,
        "Key": key,
        "Issue": issue,
        "Expected": expected,
        "Found": found,
        "Page": page
    })

def write_errors_sheet(wb, errors: list[dict]):
    if "Errors" in wb.sheetnames:
        del wb["Errors"]
    ws = wb.create_sheet("Errors")

    headers = ["Sheet", "Row", "Key", "Issue", "Expected", "Found", "Page"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    for r, rec in enumerate(errors, start=2):
        ws.cell(row=r, column=1, value=rec.get("Sheet"))
        ws.cell(row=r, column=2, value=rec.get("Row"))
        ws.cell(row=r, column=3, value=rec.get("Key"))
        ws.cell(row=r, column=4, value=rec.get("Issue"))
        ws.cell(row=r, column=5, value=rec.get("Expected"))
        ws.cell(row=r, column=6, value=rec.get("Found"))
        ws.cell(row=r, column=7, value=rec.get("Page"))

    _format_errors_ws(ws)

def load_alias_map_from_json(json_path: str) -> dict[str, str]:
    """
    Load JSON mapping Canonical -> [aliases] and return flat map: normalized(alias) -> Canonical.
    Also maps each Canonical to itself so already-clean names pass through.
    """
    alias_map: dict[str, str] = {}
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        for canonical, aliases in data.items():
            canon = str(canonical).strip()
            if not isinstance(aliases, list):
                aliases = [aliases]
            for alias in aliases:
                alias_key = normalize_label_simple(alias)
                if alias_key:
                    alias_map[alias_key] = canon
            alias_map.setdefault(normalize_label_simple(canon), canon)
    except Exception as e:
        print(f"[WARN] Could not load alias JSON: {e}")
    return alias_map

def resolve_clean_category(label: str, alias_map: dict[str, str]) -> str:
    """Return canonical name if label is an alias; otherwise return the trimmed label."""
    if not label:
        return ""
    key = normalize_label_simple(label)
    return alias_map.get(key, str(label).strip())

def _ws_header_index_map(ws) -> dict[str, int]:
    """
    Build a header->column-index (1-based) map from row 1 of an openpyxl worksheet.
    Keys are lowercase stripped header names.
    """
    header_map = {}
    for c in range(1, ws.max_column + 1):
        hdr = ws.cell(row=1, column=c).value
        if hdr is not None:
            header_map[str(hdr).strip().lower()] = c
    return header_map

def _first_blank_or_new_named_notes_col(ws, preferred_name="Notes_OrderCheck") -> int:
    """
    Find the first empty column and give it a header that doesn't collide with existing headers.
    Returns the column index (1-based).
    """
    existing_headers = set()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            existing_headers.add(str(v).strip())

    col = first_empty_col(ws)
    label = preferred_name
    i = 2
    while label in existing_headers:
        label = f"{preferred_name}_{i}"
        i += 1

    ws.cell(row=1, column=col, value=label)
    return col

def annotate_cell(ws, row: int, col: int, note: str):
    """Write/append a note string into ws[row, col]."""
    cell = ws.cell(row=row, column=col)
    existing = str(cell.value) if cell.value is not None else ""
    cell.value = (existing + ("\n" if existing else "") + note).strip()

def normalize_label_simple(s: str) -> str:
    """Lowercase + collapse spaces (including line breaks); keep symbols like '&'."""
    if not s:
        return ""
    # Replace line breaks and other whitespace with spaces, then collapse
    text = str(s).replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text.strip().lower())

def _norm_company_name(s: str) -> str:
    """
    Normalize company names for comparison.
    """
    if not s:
        return ""
    x = normalize_label_simple(s)
    x = re.sub(r'\b(inc|inc\.|llc|co|co\.|corp|corp\.|corporation|company)\b', '', x)
    x = re.sub(r'[^a-z0-9 &]+', ' ', x)
    x = re.sub(r'\s+', ' ', x).strip()
    return x

def _safe_int_local(x):
    try:
        return int(float(x))
    except Exception:
        return None

def compare_orders_with_csv(target_xlsx_path: str, ref_excel_path: str):
    """
    Company-based matching between Listings_Split and expected-order CSV.
    """
    if not os.path.isfile(ref_excel_path):
        print(f"[WARN] CSV not found: {ref_excel_path}. Skipping order comparison.")
        return

    ref_file = pd.read_csv(ref_excel_path)
    if "Category" not in ref_file.columns or "BPR Position" not in ref_file.columns:
        raise KeyError("CSV must contain 'Category' and 'BPR Position' columns.")
    if "PublishedName" not in ref_file.columns:
        raise KeyError("CSV must contain 'PublishedName' (company name) column.")

    exp_map: Dict[Tuple[str, str], int] = {}
    for _, r in ref_file.iterrows():
        cat_raw = str(r.get("Category", "")).strip()
        co_raw  = str(r.get("PublishedName", "")).strip()
        pos     = _safe_int_local(r.get("BPR Position"))
        if not cat_raw or not co_raw or pos is None:
            continue
        cat_norm = normalize_label_simple(cat_raw)
        co_norm  = _norm_company_name(co_raw)
        exp_map.setdefault((cat_norm, co_norm), pos)

    wb = load_workbook(target_xlsx_path)
    if "Listings_Split" not in wb.sheetnames:
        print("[WARN] 'Listings_Split' not found; skipping order comparison.")
        return
    ws_lsplit = wb["Listings_Split"]

    hmap = _ws_header_index_map(ws_lsplit)
    col_cat   = hmap.get("category")
    col_order = hmap.get("company order")
    col_page  = hmap.get("page")
    if not col_cat or not col_order:
        raise KeyError("'Listings_Split' needs columns 'Category' and 'Company Order'.")

    ls_company_col = None
    for k in ["company", "company name", "company_name", "name"]:
        if k in hmap:
            ls_company_col = hmap[k]
            break
    ratings_col = hmap.get("ratings table")

    notes_col = _first_blank_or_new_named_notes_col(ws_lsplit, "Notes_OrderCheck")

    def _page_at_row(r_idx: int) -> Optional[int]:
        if not col_page:
            return None
        val = ws_lsplit.cell(row=r_idx, column=col_page).value
        try:
            return int(val)
        except Exception:
            return None

    pdf_map: Dict[Tuple[str, str], Tuple[int, Optional[int], str, Optional[str]]] = {}
    for r in range(2, ws_lsplit.max_row + 1):
        cat_val = ws_lsplit.cell(row=r, column=col_cat).value
        pos_val = ws_lsplit.cell(row=r, column=col_order).value
        if cat_val is None:
            continue
        cat_disp = str(cat_val).strip()
        pos = _safe_int_local(pos_val)
        if pos is None:
            continue

        comp_disp = None
        if ls_company_col:
            v = ws_lsplit.cell(row=r, column=ls_company_col).value
            comp_disp = str(v).strip() if v is not None else None
        if not comp_disp and ratings_col:
            v = ws_lsplit.cell(row=r, column=ratings_col).value
            comp_disp = _derive_company_from_ratings_text(v)

        if not comp_disp:
            continue

        cat_norm = normalize_label_simple(cat_disp)
        co_norm  = _norm_company_name(comp_disp)
        pdf_map.setdefault((cat_norm, co_norm), (r, pos, cat_disp, comp_disp))

    exp_keys = set(exp_map.keys())
    pdf_keys = set(pdf_map.keys())

    missing_from_pdf   = sorted(exp_keys - pdf_keys)
    missing_from_input = sorted(pdf_keys - exp_keys)

    order_mismatch: List[Tuple[int, str, str, int, int]] = []
    for key in sorted(exp_keys & pdf_keys):
        exp_pos = exp_map[key]
        row_idx, act_pos, cat_disp, comp_disp = pdf_map[key]
        if act_pos is None:
            continue
        if exp_pos != act_pos:
            order_mismatch.append((row_idx, cat_disp, comp_disp or "", exp_pos, act_pos))

    for (cat_norm, co_norm) in missing_from_input:
        row_idx, _, _, comp_disp = pdf_map.get((cat_norm, co_norm), (None, None, None, None))
        if row_idx:
            annotate_cell(ws_lsplit, row_idx, notes_col, "Missing from BBB")

    for (row_idx, cat_disp, comp_disp, exp_pos, act_pos) in order_mismatch:
        annotate_cell(
            ws_lsplit,
            row_idx,
            notes_col,
            f"Order is not as expected (Expected position: {exp_pos}; Found position: {act_pos})",
        )

    new_errors: List[Dict[str, Any]] = []

    def _key_str(cat_norm, co_norm):
        try:
            cat_disp_csv = ref_file.loc[
                ref_file["Category"].apply(lambda x: normalize_label_simple(str(x))) == cat_norm, "Category"
            ].iloc[0]
        except Exception:
            cat_disp_csv = cat_norm

        comp_disp_pdf = None
        if (cat_norm, co_norm) in pdf_map:
            comp_disp_pdf = pdf_map[(cat_norm, co_norm)][3]
        if not comp_disp_pdf:
            try:
                comp_disp_csv = ref_file.loc[
                    ref_file["PublishedName"].apply(lambda x: _norm_company_name(str(x))) == co_norm,
                    "PublishedName"
                ].iloc[0]
            except Exception:
                comp_disp_csv = co_norm
            comp_disp_out = comp_disp_csv
        else:
            comp_disp_out = comp_disp_pdf
        return f"{cat_disp_csv} | {comp_disp_out}"

    for (cat_norm, co_norm) in missing_from_pdf:
        key = _key_str(cat_norm, co_norm)
        append_error(new_errors, "Listings_Split", 0, key, "missing from PDF file",
                     "Present", "Missing", page=None)

    for (cat_norm, co_norm) in missing_from_input:
        row_idx, _, _, _ = pdf_map.get((cat_norm, co_norm), (0, None, "", ""))
        key = _key_str(cat_norm, co_norm)
        append_error(
            new_errors,
            "Listings_Split",
            int(row_idx or 0),
            key,
            "missing from BBB",
            "In CSV",
            "Not in CSV",
            page=_page_at_row(int(row_idx or 0)) if row_idx else None,
        )

    for (row_idx, cat_disp, comp_disp, exp_pos, act_pos) in order_mismatch:
        key = f"{cat_disp} | {comp_disp}"
        append_error(
            new_errors,
            "Listings_Split",
            int(row_idx),
            key,
            "order is not as expected",
            exp_pos,
            act_pos,
            page=_page_at_row(int(row_idx)),
        )

    wb.save(target_xlsx_path)

    try:
        xls = pd.ExcelFile(target_xlsx_path, engine="openpyxl")
        try:
            existing_err = pd.read_excel(xls, sheet_name="Errors")
        except ValueError:
            existing_err = pd.DataFrame(columns=["Sheet", "Row", "Key", "Issue", "Expected", "Found", "Page"])
    except Exception as e:
        print(f"[WARN] Could not read existing Errors sheet: {e}")
        existing_err = pd.DataFrame(columns=["Sheet", "Row", "Key", "Issue", "Found", "Page"])

    cols = ["Sheet", "Row", "Key", "Issue", "Expected", "Found", "Page"]
    for c in cols:
        if c not in existing_err.columns:
            existing_err[c] = None

    df_new = pd.DataFrame(new_errors, columns=cols) if new_errors else pd.DataFrame(columns=cols)
    combined = pd.concat([existing_err[cols], df_new[cols]], ignore_index=True)

    with pd.ExcelWriter(target_xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        combined.to_excel(writer, sheet_name="Errors", index=False)

    wb2 = load_workbook(target_xlsx_path)
    if "Errors" in wb2.sheetnames:
        _format_errors_ws(wb2["Errors"])
        wb2.save(target_xlsx_path)

    print(
        f"[OK] Company-based order comparison complete. "
        f"New errors: {len(df_new)} | Missing from PDF: {len(missing_from_pdf)} | "
        f"Missing from input: {len(missing_from_input)} | Mismatches: {len(order_mismatch)}"
    )

def add_errors_column_to_listings_split(xlsx_path: str):
    """
    Create/overwrite 'ERRORS' column in Listings_Split by combining 'Notes' and 'Notes_OrderCheck' (if present).
    """
    wb = load_workbook(xlsx_path)
    if "Listings_Split" not in wb.sheetnames:
        wb.save(xlsx_path)
        return

    ws = wb["Listings_Split"]
    ws.freeze_panes = "A2"

    hdr_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            hdr_map[str(v).strip().lower()] = c

    col_notes = hdr_map.get("notes")
    col_notes_order = None
    for name, idx in hdr_map.items():
        if name == "notes_ordercheck" or name.startswith("notes_ordercheck_"):
            col_notes_order = idx
            break

    col_errors = hdr_map.get("errors")
    if not col_errors:
        col_errors = first_empty_col(ws)
        ws.cell(row=1, column=col_errors, value="ERRORS")

    for r in range(2, ws.max_row + 1):
        parts = []
        if col_notes:
            val = ws.cell(row=r, column=col_notes).value
            if val and str(val).strip():
                parts.append(str(val).strip())
        if col_notes_order:
            val2 = ws.cell(row=r, column=col_notes_order).value
            if val2 and str(val2).strip():
                parts.append(str(val2).strip())
        ws.cell(row=r, column=col_errors, value="\n".join(parts) if parts else "")

    err_col_letter = get_column_letter(col_errors)
    ws.column_dimensions[err_col_letter].width = 80
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_errors)
        if isinstance(cell.value, str):
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(xlsx_path)

# --------------------------------------------------------------------
# NEW: Streamlit/file-based entry point for the *full* BPRproofing run
# --------------------------------------------------------------------
def run_bprproofing_from_paths(
    pdf_path: str,
    ref_excel_path: str,
    alias_json_path: Optional[str] = None,
) -> str:
    """
    Streamlit-friendly, file-based entry point.

    - pdf_path: path to the GB PDF
    - ref_excel_path: path to the expected-order CSV/Excel (for compare_orders_with_csv)
    - alias_json_path: optional path to category_aliases.json
      If None, will look for 'category_aliases.json' in the same folder as this script.

    Returns:
        Path to the final updated workbook (same folder as the PDF).
    """
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if not os.path.isfile(ref_excel_path):
        raise FileNotFoundError(f"Expected-order CSV not found: {ref_excel_path}")

    print(f"[BPRproofing] Using PDF: {pdf_path}")
    print(f"[BPRproofing] Using expected-order file: {ref_excel_path}")

    pages = extract_pdf_text(pdf_path)
    workbook_path = save_to_excel(pdf_path, pages)

    try:
        df_pages = pd.read_excel(workbook_path, sheet_name="Pages")
    except Exception as e:
        raise RuntimeError(f"Couldn't read 'Pages' sheet: {e}")

    df_toc, df_listings, df_profiles, used_col = build_tabs_keep_rows(df_pages)
    workbook_path = write_into_existing_workbook(workbook_path, df_toc, df_listings, df_profiles)

    # ---- TOC Review step ----
    try:
        wb = load_workbook(workbook_path)
        if "TOC" in wb.sheetnames:
            ws_toc = wb["TOC"]
            b2_raw = ws_toc["B2"].value
            b3_raw = ws_toc["B3"].value
            front_raw = normalize_text(str(b2_raw) if b2_raw is not None else "")
            back_raw  = normalize_text(str(b3_raw) if b3_raw is not None else "")

            front_clean = strip_before_toc(front_raw)
            ws_toc["B2"].value = front_clean

            front_pairs = parse_pairs_split_on_numbers(front_clean)
            back_pairs  = parse_pairs_split_on_numbers(back_raw)

            write_split_sheet(wb, front_pairs, back_pairs)
            wb.save(workbook_path)

            print(f"[OK] TOC Review written. Front pairs: {len(front_pairs)} | Back pairs: {len(back_pairs)}")
        else:
            print("[WARN] 'TOC' sheet not found; skipping TOC Review.")
    except PermissionError:
        raise PermissionError("Close the workbook in Excel and try again (TOC Review stage).")
    except Exception as e:
        print(f"[WARN] TOC Review step skipped due to error: {e}")

    # ---- Listings_Split step ----
    try:
        df_listings_sheet = pd.read_excel(workbook_path, sheet_name="Listings")
        if "text" not in df_listings_sheet.columns:
            print("[ERROR] 'Listings' sheet must contain a 'text' column. Skipping Listings_Split.")
        else:
            if "page" not in df_listings_sheet.columns:
                df_listings_sheet["page"] = None

            out_records: List[Dict[str, Any]] = []
            for _, row in df_listings_sheet.iterrows():
                out_records.extend(process_listing_row(row))
            df_out = pd.DataFrame(out_records)

            with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                df_out.to_excel(xw, index=False, sheet_name="Listings_Split")

            print(f"[OK] Listings_Split written to: {workbook_path}")
            print(f"Rows: {len(df_out)} | Unique source rows: {len(df_listings_sheet)}")
    except PermissionError:
        raise PermissionError("Close the workbook in Excel and try again (Listings_Split stage).")
    except Exception as e:
        print(f"[WARN] Listings_Split step skipped due to error: {e}")

    # ---- Validation + Errors tab ----
    try:
        wb = load_workbook(workbook_path)

        # Alias JSON path: default to file next to this script if not provided
        if alias_json_path is None:
            guess = os.path.join(os.path.dirname(__file__), "category_aliases.json")
            alias_json_path = guess if os.path.exists(guess) else None

        if alias_json_path and os.path.exists(alias_json_path):
            alias_map = load_alias_map_from_json(alias_json_path)
            print(f"[INFO] Loaded {len(alias_map)} alias entries from {alias_json_path}")
        else:
            alias_map = {}
            print("[WARN] No alias JSON found; continuing without aliases.")

        errors: list[dict] = []

        # ===== A) TOC Review: Front vs Back numbers must match =====
        if "TOC Review" in wb.sheetnames:
            ws_tsplit = wb["TOC Review"]
            notes_col_toc = first_empty_col(ws_tsplit)
            ws_tsplit.cell(row=1, column=notes_col_toc, value="Notes")

            col_front_num = 2  # B
            col_back_num  = 4  # D

            for r in range(2, ws_tsplit.max_row + 1):
                front_num = ws_tsplit.cell(row=r, column=col_front_num).value
                back_num  = ws_tsplit.cell(row=r, column=col_back_num).value

                try:
                    front_num_int = int(front_num) if front_num not in (None, "") else None
                except Exception:
                    front_num_int = None
                try:
                    back_num_int = int(back_num) if back_num not in (None, "") else None
                except Exception:
                    back_num_int = None

                issue = None
                if front_num_int is None and back_num_int is not None:
                    issue = "Front TOC # missing; Back TOC # present"
                elif front_num_int is not None and back_num_int is None:
                    issue = "Back TOC # missing; Front TOC # present"
                elif (front_num_int is not None and back_num_int is not None and front_num_int != back_num_int):
                    issue = "Front/Back TOC # mismatch"

                if issue:
                    key_val = ws_tsplit.cell(row=r, column=1).value
                    note = f"[ERROR] {issue} (Front: {front_num_int}, Back: {back_num_int})"
                    annotate_cell(ws_tsplit, r, notes_col_toc, note)
                    append_error(
                        errors, "TOC Review", r, str(key_val or ""),
                        issue, str(front_num_int), str(back_num_int),
                        page=front_num_int
                    )
        else:
            print("[WARN] 'TOC Review' not found; skipping TOC front/back comparison.")

        # ===== B) Cross-check TOC Review vs Listings_Split =====
        if "TOC Review" in wb.sheetnames and "Listings_Split" in wb.sheetnames:
            ws_tsplit = wb["TOC Review"]
            ws_lsplit = wb["Listings_Split"]

            notes_col_list = first_empty_col(ws_lsplit)
            ws_lsplit.cell(row=1, column=notes_col_list, value="Notes")

            header_idx = {}
            for c in range(1, ws_lsplit.max_column + 1):
                hdr = ws_lsplit.cell(row=1, column=c).value
                if hdr:
                    header_idx[str(hdr).strip().lower()] = c
            col_category = header_idx.get("category", 3)
            col_page     = header_idx.get("page", None)
            col_content  = header_idx.get("content", None)

            toc_map: dict[str, int | None] = {}
            for r in range(2, ws_tsplit.max_row + 1):
                front_label = ws_tsplit.cell(row=r, column=1).value
                back_label  = ws_tsplit.cell(row=r, column=3).value
                chosen_label = front_label if (front_label and str(front_label).strip()) else back_label
                if not chosen_label:
                    continue
                clean = resolve_clean_category(chosen_label, alias_map)
                front_num = ws_tsplit.cell(row=r, column=2).value
                try:
                    page_int = int(front_num) if front_num not in (None, "") else None
                except Exception:
                    page_int = None
                if clean:
                    toc_map[clean] = page_int

            if col_page is None:
                print("[WARN] 'page' column not found in Listings_Split; skipping page check.")
            else:
                for r in range(2, ws_lsplit.max_row + 1):
                    cat_val = ws_lsplit.cell(row=r, column=col_category).value
                    page_val = ws_lsplit.cell(row=r, column=col_page).value

                    cat_key = (str(cat_val).strip() if cat_val is not None else "")
                    try:
                        page_int = int(page_val) if page_val not in (None, "") else None
                    except Exception:
                        page_int = None

                    if not cat_key:
                        continue

                    if cat_key not in toc_map:
                        note = f"[ERROR] Category not found in TOC Review (using aliases): '{cat_key}'"
                        annotate_cell(ws_lsplit, r, notes_col_list, note)
                        append_error(
                            errors, "Listings_Split", r, cat_key,
                            "Category missing in TOC Review (Clean/Alias)",
                            "Present in TOC Review (Clean)", "Missing",
                            page=page_int
                        )
                    else:
                        expected_page = toc_map[cat_key]
                        if page_int is None or expected_page is None or page_int != expected_page:
                            note = f"[ERROR] Page mismatch for '{cat_key}' (Expected: {expected_page}, Found: {page_int})"
                            annotate_cell(ws_lsplit, r, notes_col_list, note)
                            append_error(
                                errors, "Listings_Split", r, cat_key,
                                "Page mismatch vs TOC Review (Clean/Alias)",
                                str(expected_page), str(page_int),
                                page=page_int
                            )

            # Strong blank detection for Content
            if col_content is None:
                col_content = header_idx.get("content")

            import unicodedata  # noqa: F401

            HIDDEN_WHITESPACE = {
                "\xa0",
                "\u2000", "\u2001", "\u2002", "\u2003", "\u2004", "\u2005",
                "\u2006", "\u2007", "\u2008", "\u2009", "\u200a", "\u202f",
                "\u205f", "\u3000",
                "\u200b", "\u200c", "\u200d",
                "\ufeff",
            }
            FAKE_BLANKS = {"na", "n/a", "none", "null", "nan", "-", "–", "—", "…", "n\\a"}

            def _normalize_ws_and_strip(s: str) -> str:
                if s is None:
                    return ""
                for ch in HIDDEN_WHITESPACE:
                    s = s.replace(ch, " ")
                s = " ".join(s.split())
                return s.strip()

            def _is_blank_content(val) -> bool:
                if val is None:
                    return True
                s = _normalize_ws_and_strip(str(val))
                if s == "":
                    return True
                if s.lower() in FAKE_BLANKS:
                    return True
                if s in {"•", "·", "■", ".", "..."}:
                    return True
                return False

            if col_content is None:
                annotate_cell(ws_lsplit, 2, notes_col_list,
                              "[WARN] No 'Content' column found after normalization. Skipping blank-content checks.")
            else:
                flagged_count = 0
                for r in range(2, ws_lsplit.max_row + 1):
                    cat_val = ws_lsplit.cell(row=r, column=col_category).value if col_category else None
                    cat_key = (str(cat_val).strip() if cat_val is not None else "")
                    page_val = ws_lsplit.cell(row=r, column=col_page).value if col_page else None
                    try:
                        page_int_for_err = int(page_val) if page_val not in (None, "") else None
                    except Exception:
                        page_int_for_err = None

                    content_val = ws_lsplit.cell(row=r, column=col_content).value
                    if _is_blank_content(content_val):
                        annotate_cell(ws_lsplit, r, notes_col_list, "[ERROR] missing category content")
                        append_error(
                            errors,
                            sheet="Listings_Split",
                            row=r,
                            key=cat_key,
                            issue="missing category content",
                            expected="Non-empty",
                            found="Empty",
                            page=page_int_for_err
                        )
                        flagged_count += 1

                print(f"[INFO] Blank-content check complete. Flagged rows: {flagged_count}")

            listings_cats = set()
            for r in range(2, ws_lsplit.max_row + 1):
                cv = ws_lsplit.cell(row=r, column=col_category).value
                if cv and str(cv).strip():
                    # Normalize the same way as TOC categories for comparison
                    listings_cats.add(normalize_label_simple(str(cv).strip()))

            try:
                notes_col_toc
            except NameError:
                notes_col_toc = first_empty_col(ws_tsplit)
                ws_tsplit.cell(row=1, column=notes_col_toc, value="Notes")

            for rr in range(2, ws_tsplit.max_row + 1):
                fl = ws_tsplit.cell(row=rr, column=1).value
                bl = ws_tsplit.cell(row=rr, column=3).value
                chosen = fl if (fl and str(fl).strip()) else bl
                if not chosen:
                    continue
                clean_label = resolve_clean_category(chosen, alias_map)
                # Normalize for comparison
                clean_label_normalized = normalize_label_simple(clean_label)
                if clean_label_normalized in listings_cats:
                    continue
                # Check if it's in the ignore list (normalize both sides)
                if any(clean_label_normalized == normalize_label_simple(ignore_cat) for ignore_cat in IGNORE_TOC_ONLY_CATEGORIES):
                    print(f"[INFO] Skipping ignored category: '{clean_label}'")
                    continue

                front_num_cell = ws_tsplit.cell(row=rr, column=2).value
                try:
                    front_num_int2 = int(front_num_cell) if front_num_cell not in (None, "") else None
                except Exception:
                    front_num_int2 = None

                note = f"[ERROR] No matching category in Listings_Split for Clean '{clean_label}'"
                annotate_cell(ws_tsplit, rr, notes_col_toc, note)
                append_error(
                    errors, "TOC Review", rr, clean_label,
                    "Category missing in Listings_Split",
                    "Present in Listings_Split.Category", "Missing",
                    page=front_num_int2
                )
        else:
            print("[WARN] Missing 'TOC Review' or 'Listings_Split'; skipping cross-checks.")

        write_errors_sheet(wb, errors)
        wb.save(workbook_path)
        print(f"[OK] Validation complete. Errors logged to 'Errors' tab. Total errors: {len(errors)}")

    except PermissionError:
        raise PermissionError("Close the workbook in Excel and try again (validation stage).")
    except Exception as e:
        print(f"[WARN] Validation stage skipped due to error: {e}")

    # ---- CSV comparison ----
    try:
        compare_orders_with_csv(workbook_path, ref_excel_path)
    except PermissionError:
        raise PermissionError("Close the workbook in Excel and try again (CSV comparison stage).")
    except Exception as e:
        print(f"[WARN] CSV comparison stage skipped due to error: {e}")

    # ---- ERRORS column in Listings_Split ----
    try:
        add_errors_column_to_listings_split(workbook_path)
        print("[OK] 'ERRORS' column added to Listings_Split.")
    except PermissionError:
        raise PermissionError("Close the workbook in Excel and try again (adding ERRORS column).")
    except Exception as e:
        print(f"[WARN] Adding ERRORS column skipped due to error: {e}")

    # ---- Delete original sheets ----
    try:
        sheets_to_delete = ["Pages", "Summary", "TOC", "Listings", "Profiles"]
        delete_sheets(workbook_path, sheets_to_delete)
        print(f"[OK] Deleted {', '.join(sheets_to_delete)} from {workbook_path}")
    except PermissionError:
        raise PermissionError("Close the workbook in Excel and try again (deleting sheets).")
    except Exception as e:
        print(f"[WARN] Could not delete sheets: {e}")

    print(f"[OK] Updated workbook: {workbook_path}")
    print(f"  Used column for Listings search: {used_col}")
    print(f"  TOC rows: {len(df_toc)} | Listings rows: {len(df_listings)} | Profiles rows: {len(df_profiles)}")

    return workbook_path

def main(
    pdf_path: str,
    ref_excel_path: str,
    alias_json_path: Optional[str] = None,
) -> str:
    """
    Entry point used by bpr_pipeline.run_bprproofing_inprocess(...).

    bpr_pipeline does:
        sl_bprproofing.main(pdf_path=..., ref_excel_path=...)

    This just forwards to run_bprproofing_from_paths, which does the full
    file-based BPR proofing run and returns the final workbook path.
    """
    return run_bprproofing_from_paths(
        pdf_path=pdf_path,
        ref_excel_path=ref_excel_path,
        alias_json_path=alias_json_path,
    )

# --------------------------------------------------------------------
# Existing bytes-based Streamlit helper (unchanged)
# --------------------------------------------------------------------
def run_pipeline(pdf_bytes: bytes,
                 expected_order_df: Optional[pd.DataFrame] = None) -> Dict[str, pd.DataFrame]:
    """
    Streamlit-friendly entry point (partial):
    - Takes PDF bytes
    - Builds Pages, TOC, Listings, Profiles, TOC Review, Listings_Split
    - Returns DataFrames (no CSV comparison / validation)
    """
    tmpdir = tempfile.mkdtemp()
    pdf_path = os.path.join(tmpdir, "input.pdf")
    with open(pdf_path, "wb") as f:
        f.write(pdf_bytes)

    pages = extract_pdf_text(pdf_path)
    xlsx_path = save_to_excel(pdf_path, pages)

    try:
        df_pages = pd.read_excel(xlsx_path, sheet_name="Pages")
    except Exception as e:
        print(f"[ERROR] Couldn't read 'Pages' sheet in run_pipeline: {e}")
        df_pages = pd.DataFrame(columns=["page", "text"])

    try:
        df_toc, df_listings_raw, df_profiles, used_col = build_tabs_keep_rows(df_pages)
    except Exception as e:
        print(f"[ERROR] Parsing failed in run_pipeline: {e}")
        df_toc = pd.DataFrame()
        df_listings_raw = pd.DataFrame()
        df_profiles = pd.DataFrame()
        used_col = "<unknown>"

    try:
        out_path = write_into_existing_workbook(xlsx_path, df_toc, df_listings_raw, df_profiles)
    except Exception as e:
        print(f"[ERROR] Failed to write TOC/Listings/Profiles in run_pipeline: {e}")
        out_path = xlsx_path

    df_toc_review = pd.DataFrame()
    try:
        wb = load_workbook(out_path)
        if "TOC" in wb.sheetnames:
            ws_toc = wb["TOC"]
            b2_raw = ws_toc["B2"].value
            b3_raw = ws_toc["B3"].value
            front_raw = normalize_text(str(b2_raw) if b2_raw is not None else "")
            back_raw  = normalize_text(str(b3_raw) if b3_raw is not None else "")

            front_clean = strip_before_toc(front_raw)
            ws_toc["B2"].value = front_clean

            front_pairs = parse_pairs_split_on_numbers(front_clean)
            back_pairs  = parse_pairs_split_on_numbers(back_raw)

            write_split_sheet(wb, front_pairs, back_pairs)
            wb.save(out_path)

            try:
                df_toc_review = pd.read_excel(out_path, sheet_name="TOC Review")
            except Exception as e:
                print(f"[WARN] Could not re-read 'TOC Review' as DataFrame: {e}")
                df_toc_review = pd.DataFrame()
        else:
            print("[WARN] 'TOC' sheet not found; skipping TOC Review in run_pipeline.")
            wb.close()
    except Exception as e:
        print(f"[WARN] TOC Review step skipped in run_pipeline due to error: {e}")
        df_toc_review = pd.DataFrame()

    df_listings_split = pd.DataFrame()
    try:
        df_listings_sheet = pd.read_excel(out_path, sheet_name="Listings")
    except Exception as e:
        print(f"[ERROR] Couldn't read 'Listings' sheet in run_pipeline: {e}")
        df_listings_sheet = pd.DataFrame()

    try:
        if "text" in df_listings_sheet.columns:
            if "page" not in df_listings_sheet.columns:
                df_listings_sheet["page"] = None

            out_records: List[Dict[str, Any]] = []
            for _, row in df_listings_sheet.iterrows():
                out_records.extend(process_listing_row(row))
            df_listings_split = pd.DataFrame(out_records)

            with pd.ExcelWriter(out_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                df_listings_split.to_excel(xw, index=False, sheet_name="Listings_Split")

            print(f"[OK] Listings_Split created in run_pipeline. Rows: {len(df_listings_split)}")
        else:
            print("[WARN] 'Listings' sheet missing 'text' column; no Listings_Split built in run_pipeline.")
    except Exception as e:
        print(f"[WARN] Listings_Split step skipped in run_pipeline due to error: {e}")
        df_listings_split = pd.DataFrame()

    results: Dict[str, pd.DataFrame] = {}
    results["Listings_Split"] = df_listings_split
    results["Errors"] = pd.DataFrame()
    results["TOC Presence Check"] = pd.DataFrame()
    results["TOC Review"] = df_toc_review
    results["Profiles"] = df_profiles
    results["Listings"] = df_listings_raw
    results["Pages"] = df_pages

    return results
