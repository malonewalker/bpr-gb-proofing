#!/usr/bin/env python3
"""
Streamlit version of the BPR Unified Runner

What this version does:
- Uses one validated Profiles sheet named exactly "Profiles".
- Adds "TOC Review" if sl_proofing returns it.
- Builds a single Errors tab by combining:
    • Errors from sl_bprproofing
    • Errors from sl_newvalidate
    • Profiles-synthesized rows from Notes_Internal + Notes_Compare (always included if present)
- Cleans placeholder lines like "[Internal] nan" / "[Compare] nan" and drops empty-issue rows.
- Adds a simple header AutoFilter ONLY on the Errors tab (no Excel Tables).
"""

from __future__ import annotations

import io
import os
import re
import traceback
from datetime import datetime
from typing import Optional, List, Dict, Any

import pandas as pd
import streamlit as st

# --- Local modules (your updated Streamlit-friendly scripts) ---
# Adjust these imports / function names to match your actual sl_* files.
import sl_bprproofing as bpr
import sl_proofing as prof
import sl_newvalidate as nv

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------
# Config
# -----------------------------
PROFILES_VALIDATED_NAME = "Profiles"
KEEP_RAW_PROFILES = False
RAW_PROFILES_NAME = "Profiles_Raw"

_EMPTY_TOKENS = {"", "nan", "none", "null", "n/a", "na", "-"}


# -----------------------------
# Helpers: text normalization & cleaning
# -----------------------------
def _norm_text(v: object) -> str:
    s = ("" if v is None else str(v)).strip()
    return "" if s.lower() in _EMPTY_TOKENS else s


def _clean_issue_text(issue: object) -> str:
    """Remove 'nan'-style placeholders and empty labeled lines."""
    s = ("" if issue is None else str(issue))

    # Collapse whitespace/newlines
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n+", "\n", s).strip()

    # Drop lines that are just placeholders or label+placeholder
    lines = []
    for line in s.split("\n"):
        line = line.strip()
        if line.lower() in _EMPTY_TOKENS:
            continue
        if re.fullmatch(r"\[(Internal|Compare)\]\s*(nan|none|null|n/?a|-)?\s*", line, flags=re.I):
            continue
        if re.fullmatch(r"\[(Internal|Compare)\]\s*", line, flags=re.I):
            continue
        lines.append(line)

    s = "\n".join(lines).strip()
    return s


# -----------------------------
# Errors from Profiles (DataFrame version)
# -----------------------------
def _errors_from_profiles_df(df_in: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Build an Errors-style DataFrame from a Profiles DataFrame:

    Columns: Sheet, Row, Key, Issue, Expected, Found, Page

    - Prefers a ready-made ERRORS column.
    - Otherwise combines Notes_Internal + Notes_Compare into Issue.
    """
    if df_in is None or not isinstance(df_in, pd.DataFrame) or df_in.empty:
        return None

    df = df_in.copy()
    df.columns = [str(c).strip() for c in df.columns]

    def find_col(cands):
        lcands = [x.lower().strip() for x in cands]
        for c in df.columns:
            if c.lower().strip() in lcands:
                return c
        return None

    # Prefer a ready-made ERRORS column if present
    col_errors = find_col(["ERRORS", "Errors", "Error", "Issue", "Issues"])

    col_notes_i = find_col(["Notes_Internal", "Notes Internal", "Internal Notes", "Notes (Internal)", "Internal"])
    col_notes_c = find_col(["Notes_Compare", "Notes Compare", "Compare Notes", "Notes (Compare)", "Compare"])
    col_category = find_col(["Category", "Cat", "Category Name"])
    col_pubnum = find_col(
        [
            "Published Name + Number",
            "Published Name+Number",
            "Published Name Number",
            "Published Name/Number",
            "Published Name & Number",
            "Published Name and Number",
            "Published Name",
            "Name + Number",
            "Name+Number",
        ]
    )
    col_page = find_col(["Page", "Pg", "Page #", "Page Number", "PageNo"])

    # Build Issue text
    if col_errors:
        issues = df[col_errors].map(_clean_issue_text)
    else:
        def build_issue(row):
            parts = []
            if col_notes_i:
                v = _norm_text(row.get(col_notes_i, ""))
                if v:
                    parts.append(f"[Internal] {v}")
            if col_notes_c:
                v = _norm_text(row.get(col_notes_c, ""))
                if v:
                    parts.append(f"[Compare] {v}")
            return _clean_issue_text("\n".join(parts))

        issues = df.apply(build_issue, axis=1)

    mask = issues.astype(str).str.strip().astype(bool)
    if not mask.any():
        return None

    excel_rows = (df.index + 2).astype(int)

    def build_key(row):
        left = _norm_text(row.get(col_category, "")) if col_category else ""
        right = _norm_text(row.get(col_pubnum, "")) if col_pubnum else ""
        return f"{left} / {right}".strip(" /")

    out = pd.DataFrame(
        {
            "Sheet": "Profiles",
            "Row": excel_rows,
            "Key": df.apply(build_key, axis=1),
            "Issue": issues,
            "Expected": "",
            "Found": "",
            "Page": df[col_page] if col_page in df.columns else "",
        }
    )

    return out[mask].reset_index(drop=True)


# -----------------------------
# Error dedupe + formatting
# -----------------------------
def _dedupe_errors(df: pd.DataFrame) -> pd.DataFrame:
    """Light dedupe by common subset."""
    for col in ["Sheet", "Key", "Issue", "Expected", "Found", "Page"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    subset = [c for c in ["Sheet", "Row", "Key", "Issue", "Page"] if c in df.columns]
    if subset:
        return df.drop_duplicates(subset=subset, keep="first").reset_index(drop=True)
    return df


def _excel_autowidth(ws, max_width=60):
    """Autosize columns based on first line of content; cap width."""
    for col in ws.columns:
        first_cell = next((c for c in col if c is not None), None)
        if first_cell is None:
            continue
        try:
            letter = get_column_letter(
                getattr(first_cell, "column", getattr(first_cell, "col_idx", 1))
            )
        except Exception:
            continue
        maxlen = 0
        for cell in col:
            if cell is None:
                continue
            val = "" if cell.value is None else str(cell.value)
            first_line = val.split("\n", 1)[0]
            maxlen = max(maxlen, len(first_line) + (3 if "\n" in val else 0))
        ws.column_dimensions[letter].width = min(maxlen + 2, max_width)


def _write_df_sheet(wb, sheet_name: str, df: pd.DataFrame, *, add_header_filter: bool = False):
    """Create/replace sheet, write DataFrame, style header, freeze top row,
       and optionally add a simple header AutoFilter (NO Excel Table)."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    hdr_font = Font(bold=True)
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = hdr_font
        c.alignment = Alignment(vertical="top", wrap_text=True)

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    if add_header_filter and df.shape[0] >= 1 and df.shape[1] >= 1:
        last_col = get_column_letter(df.shape[1])
        last_row = df.shape[0] + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    _excel_autowidth(ws)
    return ws


# -----------------------------
# Combine Errors from all sources
# -----------------------------
def build_combined_errors(
    bpr_errors: Optional[pd.DataFrame],
    nv_errors: Optional[pd.DataFrame],
    profiles_validated: Optional[pd.DataFrame],
) -> pd.DataFrame:
    sources: List[pd.DataFrame] = []

    def attach(df: Optional[pd.DataFrame], label: str):
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            return
        out = df.copy()
        if "Source" not in out.columns:
            out["Source"] = label
        sources.append(out)

    attach(bpr_errors, "BPRproofing")
    attach(nv_errors, "newvalidate")

    prof_synth = _errors_from_profiles_df(profiles_validated) if profiles_validated is not None else None
    attach(prof_synth, "Profiles")

    if not sources:
        return pd.DataFrame(columns=["Sheet", "Row", "Key", "Issue", "Expected", "Found", "Page", "Source"])

    # Union of columns
    all_cols: List[str] = []
    for d in sources:
        for c in d.columns:
            if c not in all_cols:
                all_cols.append(c)

    combined = pd.concat([d.reindex(columns=all_cols) for d in sources], ignore_index=True)

    # Clean up issues
    if "Issue" in combined.columns:
        combined["Issue"] = combined["Issue"].map(_clean_issue_text)
        combined = combined[combined["Issue"].astype(str).str.strip().astype(bool)].reset_index(drop=True)

    desired = ["Sheet", "Row", "Key", "Issue", "Expected", "Found", "Page", "Source"]
    ordered = [c for c in desired if c in combined.columns] + [
        c for c in combined.columns if c not in desired
    ]
    combined = combined.reindex(columns=ordered)
    combined = _dedupe_errors(combined)

    return combined


# -----------------------------
# Build Excel bytes for download
# -----------------------------
def build_excel_bytes(
    results: Dict[str, pd.DataFrame],
    bpr_errors: Optional[pd.DataFrame],
    nv_errors: Optional[pd.DataFrame],
) -> bytes:
    """
    Writes the combined workbook into memory and returns bytes ready for download.
    - Writes main sheets with pandas.
    - Then reopens with openpyxl to add the combined Errors sheet with AutoFilter.
    - Renames a couple of tabs for final tidy.
    """
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Optionally raw Profiles
        if KEEP_RAW_PROFILES and isinstance(results.get("Profiles_Raw"), pd.DataFrame):
            results["Profiles_Raw"].to_excel(writer, index=False, sheet_name=RAW_PROFILES_NAME)

        # Validated Profiles
        if isinstance(results.get("Profiles"), pd.DataFrame):
            results["Profiles"].to_excel(writer, index=False, sheet_name=PROFILES_VALIDATED_NAME)

        # Other useful sheets
        for name in ["Listings_Split", "TOC Presence Check", "TOC Review", "Listings", "Pages"]:
            df = results.get(name)
            if isinstance(df, pd.DataFrame) and not df.empty:
                df.to_excel(writer, index=False, sheet_name=name)

    out.seek(0)
    wb = load_workbook(out)

    # Rename tabs like the old tidy_up_final_workbook
    rename_map = {
        "Listings_Split": "Cat Listings",
        "TOC Review": "TOC",
    }
    for old_name, new_name in rename_map.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name

    # Remove raw if we created it and don't want it
    if not KEEP_RAW_PROFILES and RAW_PROFILES_NAME in wb.sheetnames:
        del wb[RAW_PROFILES_NAME]

    # Combined Errors
    profiles_validated = results.get("Profiles")
    combined_errors = build_combined_errors(bpr_errors, nv_errors, profiles_validated)

    _write_df_sheet(wb, "Errors", combined_errors, add_header_filter=True)

    out2 = io.BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.getvalue()


# -----------------------------
# Wrappers over sl_* modules
# (adjust function names here if yours differ)
# -----------------------------
def run_sl_bpr(pdf_bytes: bytes, ref_file: Optional[pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Expected: return a dict of DataFrames with keys like:
      'Listings_Split', 'Listings', 'Pages', 'Profiles', 'Errors'
    Adjust this wrapper to match your actual sl_bprproofing API.
    """
    # Example assumption:
    return bpr.run_pipeline(pdf_bytes=pdf_bytes, expected_order_df=ref_file)


def run_sl_proofing(pdf_bytes: bytes, ref_file: Optional[pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Expected keys:
      'TOC Presence Check', 'TOC Review', maybe 'Errors'
    """
    return prof.run_pipeline(pdf_bytes=pdf_bytes, expected_order_df=ref_file)


def run_sl_newvalidate(
    profiles_df: pd.DataFrame, ref_file: pd.DataFrame
) -> tuple[pd.DataFrame, Optional[pd.DataFrame]]:
    """
    Expected to behave like newvalidate.run_checks but for DataFrames, e.g.:
      - returns (checked_df, errors_df)  OR
      - returns dict with 'Profiles' and 'Errors'  OR
      - returns just checked_df
    This wrapper normalizes that into (checked_df, errors_df_or_None).
    """
    result = nv.run_checks(profiles_df.copy(), ref_file)

    checked_df: pd.DataFrame
    nv_errors_df: Optional[pd.DataFrame] = None

    if isinstance(result, tuple):
        checked_df = result[0]
        if len(result) > 1 and isinstance(result[1], pd.DataFrame):
            nv_errors_df = result[1]
    elif isinstance(result, dict):
        checked_df = (
            result.get("Profiles")
            or result.get("profiles")
            or result.get("validated")
            or result.get("checked")
        )
        if not isinstance(checked_df, pd.DataFrame):
            raise RuntimeError("sl_newvalidate did not return a DataFrame for Profiles.")
        if isinstance(result.get("Errors"), pd.DataFrame):
            nv_errors_df = result["Errors"]
    else:
        checked_df = result  # assume DataFrame

    return checked_df, nv_errors_df


# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title="BPR — Book Proofing App (Streamlit)", layout="wide")
st.title("BPR — Book Proofing App")

with st.sidebar:
    st.header("Inputs")

    pdf_file = st.file_uploader("Upload Book PDF", type=["pdf"])

    # 🔁 One reference file used for BOTH Expected Order and BBB (newvalidate)
    ref_file = st.file_uploader(
        "Upload BBB",
        type=["csv", "xlsx", "xls", "txt"],
        help="Same file is used for TOC/Expected Order checks and BBB/newvalidate checks."
    )

    run_btn = st.button("Run All")


def _read_tabular(upload) -> Optional[pd.DataFrame]:
    if upload is None:
        return None
    name = upload.name.lower()
    try:
        if name.endswith(".csv") or name.endswith(".txt"):
            return pd.read_csv(upload)
        else:
            return pd.read_excel(upload)
    except Exception as e:
        st.error(f"Could not read file {upload.name}: {e}")
        return None

results: Optional[Dict[str, pd.DataFrame]] = None
bpr_errors_df: Optional[pd.DataFrame] = None
nv_errors_df: Optional[pd.DataFrame] = None

if run_btn:
    if not pdf_file:
        st.warning("Please upload a PDF.")
    elif not ref_file:
        st.warning("Please upload an Expected Order file (CSV/Excel).")

    else:
        with st.status("Processing…", expanded=False) as s:
            try:
                pdf_bytes = pdf_file.read()
                ref_file = _read_tabular(ref_file)

                if ref_file is None or ref_file is None:
                    st.stop()

                # 1) BPRproofing (Listings, Profiles, Pages, Errors)
                bpr_out = run_sl_bpr(pdf_bytes, ref_file)

                # 2) Profiles validation
                raw_profiles = bpr_out.get("Profiles")
                if raw_profiles is None or not isinstance(raw_profiles, pd.DataFrame) or raw_profiles.empty:
                    st.error("No Profiles data returned from sl_bprproofing; cannot run newvalidate.")
                    st.stop()

                checked_profiles, nv_errors_df = run_sl_newvalidate(raw_profiles, ref_file)

                # 3) TOC checks
                proof_out = run_sl_proofing(pdf_bytes, ref_file)

                # Collect results dict
                results = {}
                # Raw Profiles if you want to keep it visible
                if KEEP_RAW_PROFILES:
                    results["Profiles_Raw"] = raw_profiles
                # Validated Profiles
                results["Profiles"] = checked_profiles

                # BPR outputs
                for key in ["Listings_Split", "Listings", "Pages"]:
                    if isinstance(bpr_out.get(key), pd.DataFrame):
                        results[key] = bpr_out[key]
                if isinstance(bpr_out.get("Errors"), pd.DataFrame):
                    bpr_errors_df = bpr_out["Errors"]

                # TOC outputs
                for key in ["TOC Presence Check", "TOC Review"]:
                    if isinstance(proof_out.get(key), pd.DataFrame):
                        results[key] = proof_out[key]
                # Optional extra TOC errors
                if isinstance(proof_out.get("Errors"), pd.DataFrame):
                    # Treat these as an additional error source under "BPRproofing" or separate if you like
                    if bpr_errors_df is None:
                        bpr_errors_df = proof_out["Errors"]
                    else:
                        bpr_errors_df = pd.concat([bpr_errors_df, proof_out["Errors"]], ignore_index=True)

                s.update(label="Done.", state="complete")
            except Exception as e:
                st.error(f"Pipeline failed: {e}")
                traceback.print_exc()
                results = None

if results:
    tabs = st.tabs(
        [
            "Listings_Split",
            "Errors",
            "TOC Presence Check",
            "TOC Review",
            "Profiles",
            "Listings",
            "Pages",
        ]
    )

    with tabs[0]:
        st.subheader("Listings_Split")
        df = results.get("Listings_Split")
        if isinstance(df, pd.DataFrame):
            st.dataframe(df, use_container_width=True, height=420)
        else:
            st.info("No Listings_Split data.")

    with tabs[1]:
        st.subheader("Errors (Merged)")
        combined_errors = build_combined_errors(
            bpr_errors_df,
            nv_errors_df,
            results.get("Profiles"),
        )
        st.dataframe(combined_errors, use_container_width=True, height=420)

    with tabs[2]:
        st.subheader("TOC Presence Check (true missing only)")
        df = results.get("TOC Presence Check")
        if isinstance(df, pd.DataFrame):
            st.dataframe(df, use_container_width=True, height=420)
        else:
            st.info("No TOC Presence Check data.")
        st.caption(
            "Order-insensitive & punctuation-agnostic matching avoids false flags like "
            "‘Heating & Air Conditioning’ vs ‘Air Conditioning & Heating’."
        )

    with tabs[3]:
        st.subheader("TOC Review (Front/Back)")
        df = results.get("TOC Review")
        if isinstance(df, pd.DataFrame):
            st.dataframe(df, use_container_width=True, height=300)
        else:
            st.info("No TOC Review data.")

    with tabs[4]:
        st.subheader("Profiles (validated)")
        df = results.get("Profiles")
        if isinstance(df, pd.DataFrame):
            st.dataframe(df, use_container_width=True, height=300)
        else:
            st.info("No Profiles data.")

    with tabs[5]:
        st.subheader("Listings (raw)")
        df = results.get("Listings")
        if isinstance(df, pd.DataFrame):
            st.dataframe(df, use_container_width=True, height=300)
        else:
            st.info("No Listings data.")

    with tabs[6]:
        st.subheader("Pages (raw)")
        df = results.get("Pages")
        if isinstance(df, pd.DataFrame):
            st.dataframe(df, use_container_width=True, height=300)
        else:
            st.info("No Pages data.")

    # Download Excel
    xls_bytes = build_excel_bytes(results, bpr_errors_df, nv_errors_df)
    st.download_button(
        "Download Combined Excel",
        data=xls_bytes,
        file_name=f"BPR_Combined_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
else:
    st.info("Upload the Book PDF, upload the BBB file, and click **Run All**.")
