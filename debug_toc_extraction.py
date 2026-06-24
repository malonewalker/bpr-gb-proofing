#!/usr/bin/env python3
"""
Create a TOC debug workbook from a PDF without running the full app pipeline.

Usage:
    python debug_toc_extraction.py "path/to/book.pdf"
    python debug_toc_extraction.py "path/to/book.pdf" --out "path/to/debug.xlsx"

The workbook keeps the raw extraction tabs so TOC parser failures can be
debugged from VS Code:
    - Pages
    - Summary
    - TOC
    - Listings
    - Profiles
    - TOC Review
    - TOC Raw Text
    - TOC Debug Lines
"""

from __future__ import annotations

import argparse
import os
import re
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from sl_bprproofing import (
    build_tabs_keep_rows,
    clean_toc_line,
    extract_pdf_text,
    normalize_text,
    parse_pairs_split_on_numbers,
    strip_before_toc,
    write_into_existing_workbook,
    write_split_sheet,
)


TEXT_COL_CANDIDATES = ("text", "Text", "TEXT")


def _text_column(df: pd.DataFrame) -> str | None:
    for col in TEXT_COL_CANDIDATES:
        if col in df.columns:
            return col
    return None


def _default_out_path(pdf_path: str) -> str:
    folder = os.path.dirname(os.path.abspath(pdf_path))
    stem = os.path.splitext(os.path.basename(pdf_path))[0]
    return os.path.join(folder, f"{stem}_toc_debug.xlsx")


def _write_base_workbook(pdf_path: str, pages: list[str], out_path: str) -> None:
    combined = "\n\n".join(pages)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame({"page": range(1, len(pages) + 1), "text": pages}).to_excel(
            writer, index=False, sheet_name="Pages"
        )
        pd.DataFrame(
            [
                {
                    "pdf": pdf_path,
                    "pages": len(pages),
                    "characters_total": len(combined),
                    "method": "PyPDF2 (no OCR)",
                }
            ]
        ).to_excel(writer, index=False, sheet_name="Summary")


def _iter_raw_toc_blocks(df_toc: pd.DataFrame) -> Iterable[tuple[str, str]]:
    text_col = _text_column(df_toc)
    if text_col is None:
        return []

    blocks: list[tuple[str, str]] = []
    for idx, row in df_toc.reset_index(drop=True).iterrows():
        page = row.get("page", "")
        label = f"TOC row {idx + 1} / page {page}"
        raw = "" if pd.isna(row.get(text_col)) else str(row.get(text_col))
        blocks.append((label, raw))
    return blocks


def _line_debug_records(label: str, block: str) -> list[dict[str, object]]:
    text = normalize_text(block)
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    records: list[dict[str, object]] = []
    for line_no, raw_line in enumerate(text.split("\n"), start=1):
        cleaned = clean_toc_line(raw_line)
        parsed_pairs = parse_pairs_split_on_numbers(cleaned)
        records.append(
            {
                "Block": label,
                "Line": line_no,
                "Raw Line": raw_line,
                "Clean Line": cleaned,
                "Line Parses Alone": "; ".join(f"{cat} -> {num}" for cat, num in parsed_pairs),
            }
        )
    return records


def _build_raw_text_df(front_raw: str, back_raw: str) -> pd.DataFrame:
    front_clean = strip_before_toc(front_raw)
    back_clean = back_raw
    return pd.DataFrame(
        [
            {
                "Block": "front_raw",
                "Text": front_raw,
                "Parsed Pairs": "",
            },
            {
                "Block": "front_clean_after_table_of_contents",
                "Text": front_clean,
                "Parsed Pairs": "\n".join(
                    f"{cat} -> {num}" for cat, num in parse_pairs_split_on_numbers(front_clean)
                ),
            },
            {
                "Block": "back_raw",
                "Text": back_raw,
                "Parsed Pairs": "\n".join(
                    f"{cat} -> {num}" for cat, num in parse_pairs_split_on_numbers(back_clean)
                ),
            },
        ]
    )


def _format_workbook(path: str) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                first_line = value.split("\n", 1)[0]
                max_len = max(max_len, len(first_line))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 80)
    wb.save(path)
    wb.close()


def create_toc_debug_workbook(pdf_path: str, out_path: str | None = None) -> str:
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    out_path = out_path or _default_out_path(pdf_path)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)

    pages = extract_pdf_text(pdf_path)
    _write_base_workbook(pdf_path, pages, out_path)

    df_pages = pd.read_excel(out_path, sheet_name="Pages")
    df_toc, df_listings, df_profiles, used_col = build_tabs_keep_rows(df_pages)
    write_into_existing_workbook(out_path, df_toc, df_listings, df_profiles)

    raw_blocks = list(_iter_raw_toc_blocks(df_toc))
    front_raw = normalize_text(raw_blocks[0][1] if len(raw_blocks) >= 1 else "")
    back_raw = normalize_text(raw_blocks[1][1] if len(raw_blocks) >= 2 else "")
    front_clean = strip_before_toc(front_raw)

    front_pairs = parse_pairs_split_on_numbers(front_clean)
    back_pairs = parse_pairs_split_on_numbers(back_raw)

    wb = load_workbook(out_path)
    write_split_sheet(wb, front_pairs, back_pairs)
    wb.save(out_path)
    wb.close()

    raw_text_df = _build_raw_text_df(front_raw, back_raw)
    line_records: list[dict[str, object]] = []
    for label, raw in raw_blocks:
        line_records.extend(_line_debug_records(label, raw))

    summary_df = pd.DataFrame(
        [
            {"Metric": "PDF", "Value": pdf_path},
            {"Metric": "Output workbook", "Value": out_path},
            {"Metric": "Listings search column", "Value": used_col},
            {"Metric": "TOC rows", "Value": len(df_toc)},
            {"Metric": "Front TOC parsed pairs", "Value": len(front_pairs)},
            {"Metric": "Back TOC parsed pairs", "Value": len(back_pairs)},
        ]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="TOC Debug Summary")
        raw_text_df.to_excel(writer, index=False, sheet_name="TOC Raw Text")
        pd.DataFrame(line_records).to_excel(writer, index=False, sheet_name="TOC Debug Lines")

    _format_workbook(out_path)
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Create a raw TOC debug workbook from a PDF.")
    parser.add_argument("pdf", help="Path to the PDF to inspect.")
    parser.add_argument("--out", help="Optional output .xlsx path.")
    args = parser.parse_args()

    out_path = create_toc_debug_workbook(args.pdf, args.out)
    print(f"[OK] TOC debug workbook written: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
